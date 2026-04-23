"""
build_check_tab.py
------------------
Builds the Check Tab in a COPY of latestNAV_Reports.xlsx.

Does NOT touch the live latestNAV_Reports.xlsx (which update_morningstar.py
may be writing to). Works on a specified backup copy.

What it builds:
  1. AllClientHoldings sheet — consolidated client data with lookup key
       Col A: key = ClientID|RowNum  (e.g. "Rahul_Sharma|1")
       Col B: Client_ID
       Col C: ISIN
       Col D: Units
     Populated from all client files in the client directory.
     Source for VLOOKUP in Check Tab.

  2. Check Tab — fully formula-driven, Excel 2019 compatible
     Summary section (rows 1-11):
       Client ID dropdown (data validation from AllClientHoldings)
       Archetype dropdown (Averse / Moderate / Aggressive)
       New Cash (₹), Run Date
       Target Equity%, Defensive%, Other% (from archetype table)
       Current Equity%, Defensive%, Other% (SUMPRODUCT over holdings)
       Gap and Switch Amount
     Holdings detail (rows 13-163, 150 rows):
       # | ISIN | Scheme Name | Units | NAV | Value | Eq% | Def% | Other% |
       Eq Val | Def Val | Other Val | Weight%
     Transactions section (rows 166+):
       Written by bulk_run.py, read-only display here

  Risk type on Check Tab (cell C5): Excel **VLOOKUP** on the selected Client ID into sheet
  **Transactions** column **target_policy** (Averse | Moderate | Aggressive) when that sheet
  has rows; otherwise sheet **Client risk map** if the pack wrote it from **client_risk_pref**;
  no per-client hidden columns on the Check Tab.

EXCEL 2019 COMPATIBILITY:
  - No XLOOKUP (uses VLOOKUP / INDEX-MATCH)
  - No FILTER function (uses fixed 150-row range with VLOOKUP on key)
  - No dynamic arrays / spill

Usage:
  python build_check_tab.py
  python build_check_tab.py --master data/latestNAV_Reports_280326_0019.xlsx
  python build_check_tab.py --clients data/dummy_clients/by_client/

Default master: data/latestNAV_Reports.xlsx (live workbook with current formulas).
If that file is missing, use newest data/latestNAV_Reports_*.xlsx backup. Reading
the live file for the pack uses a temporary copy so a concurrent writer does not
share the same path.
"""

import argparse
import glob
import os
import re
import shutil
import sys
import tempfile
from collections import defaultdict

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import (
    Alignment, Border, Font, PatternFill, Side
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# Timestamped master backups: latestNAV_Reports_DDMMYY_HHMM.xlsx (basename string sort = newest)
_TS_MASTER_BACKUP_RE = re.compile(r"^latestNAV_Reports_\d{6}_\d{4}\.xlsx$", re.IGNORECASE)

# ── Paths ────────────────────────────────────────────────────────────────────
LIVE_NAV_REPORTS = os.path.join("data", "latestNAV_Reports.xlsx")
DEFAULT_CLIENTS  = os.path.join("data", "dummy_clients", "by_client")
DEFAULT_OUTPUT   = os.path.join("data", "Check.xlsx")


def default_master():
    """
    Prefer live data/latestNAV_Reports.xlsx (current Final cols, D/D+E, etc.).

    If the live file is absent, use the newest timestamped backup
    latestNAV_Reports_DDMMYY_HHMM.xlsx (basename sort). If no backups either,
    return the live path and let the caller error.
    """
    if os.path.isfile(LIVE_NAV_REPORTS):
        return LIVE_NAV_REPORTS
    pattern = os.path.join("data", "latestNAV_Reports_*.xlsx")
    matches = [p for p in glob.glob(pattern) if os.path.isfile(p)]
    ts_matches = [p for p in matches if _TS_MASTER_BACKUP_RE.match(os.path.basename(p))]
    pool = ts_matches if ts_matches else matches
    if pool:
        pick = max(pool, key=lambda p: os.path.basename(p))
        print(
            f"WARNING: {LIVE_NAV_REPORTS} not found; using newest backup {pick}",
            file=sys.stderr,
        )
        return pick
    print(
        f"WARNING: No NAV master file found; expected {LIVE_NAV_REPORTS}",
        file=sys.stderr,
    )
    return LIVE_NAV_REPORTS


def _same_nav_path(a: str, b: str) -> bool:
    """Stable path equality (Windows-friendly)."""
    na = os.path.normcase(os.path.abspath(os.path.normpath(a)))
    nb = os.path.normcase(os.path.abspath(os.path.normpath(b)))
    return na == nb


NAV_SHEET       = "latestNAV_Reports"
CHECK_SHEET     = "Check Tab"
HOLDINGS_SHEET  = "AllClientHoldings"
# Must match TRANSACTIONS_SHEET in build_mfd_pack.py (VLOOKUP in Check Tab).
TRANSACTIONS_SHEET_FOR_FORMULA = "Transactions"
CLIENT_RISK_MAP_SHEET = "Client risk map"

# Layout constants
MAX_HOLDINGS_ROWS = 150        # rows 14..163 in Check Tab
HOLDINGS_START    = 14         # first data row in Check Tab
HOLDINGS_END      = HOLDINGS_START + MAX_HOLDINGS_ROWS - 1  # 163
TXNS_START        = HOLDINGS_END + 3                         # 166

# Column indices in latestNAV_Reports (1-based)
NAV_COL_ISIN      = 5   # E
NAV_COL_NAME      = 4   # D
NAV_COL_NAV       = 7   # G
NAV_COL_FIN_EQ    = 21  # U
NAV_COL_FIN_DEBT  = 22  # V
NAV_COL_FIN_CASH  = 23  # W
NAV_COL_FIN_OTHER = 24  # X

# ── Styles ───────────────────────────────────────────────────────────────────
DARK_BLUE  = "1F4E79"
MID_BLUE   = "2E75B6"
YELLOW     = "FFFF00"
LIGHT_GREY = "F2F2F2"
WHITE      = "FFFFFF"
GREEN      = "E2EFDA"
AMBER      = "FFF2CC"
LIGHT_BLUE = "DDEEFF"

def hdr_style(fg=DARK_BLUE, bold=True, size=9):
    return {
        "font":  Font(bold=bold, color=WHITE, name="Arial", size=size),
        "fill":  PatternFill("solid", fgColor=fg),
        "align": Alignment(horizontal="center", vertical="center", wrap_text=True),
    }

def label_style():
    return {
        "font":  Font(bold=True, name="Arial", size=9),
        "align": Alignment(horizontal="left", vertical="center"),
    }

def input_style():
    return {
        "font":  Font(name="Arial", size=9, color="000080"),
        "fill":  PatternFill("solid", fgColor=YELLOW),
        "align": Alignment(horizontal="left", vertical="center"),
    }

def formula_style(fill_color=None):
    fill = PatternFill("solid", fgColor=fill_color) if fill_color else PatternFill()
    return {
        "font":  Font(name="Arial", size=9),
        "fill":  fill,
        "align": Alignment(horizontal="right", vertical="center"),
    }

def apply(cell, style_dict):
    if "font"   in style_dict: cell.font      = style_dict["font"]
    if "fill"   in style_dict: cell.fill      = style_dict["fill"]
    if "align"  in style_dict: cell.alignment = style_dict["align"]
    if "border" in style_dict: cell.border    = style_dict["border"]
    if "fmt"    in style_dict: cell.number_format = style_dict["fmt"]

def thin_border():
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)

# ── Client data loading ──────────────────────────────────────────────────────
ISIN_PAT = re.compile(r'^IN[A-Z0-9]{10}$')

def detect_isin_col(df):
    for col in df.columns:
        sample = df[col].dropna().astype(str).str.strip()
        if len(sample) and sample.apply(lambda v: bool(ISIN_PAT.match(v))).mean() >= 0.5:
            return col
    return None

def detect_units_col(df, isin_col):
    kw = ["unit", "qty", "quantity", "balance", "holding"]
    numeric = [c for c in df.columns if c != isin_col
               and pd.api.types.is_numeric_dtype(df[c])]
    for c in numeric:
        if any(k in c.lower() for k in kw):
            return c
    for c in numeric:
        vals = df[c].dropna()
        if len(vals) and (vals > 0).all():
            return c
    return None

def load_all_clients(client_dir):
    """
    Load all client files from directory.
    Returns list of (client_id, isin, units) tuples.
    """
    rows = []
    files = [f for f in sorted(os.listdir(client_dir))
             if f.lower().endswith((".xlsx", ".xls", ".csv"))]
    print(f"  Loading {len(files)} client files from {client_dir}")
    for fname in files:
        client_id = os.path.splitext(fname)[0]
        fpath = os.path.join(client_dir, fname)
        try:
            df = (pd.read_csv(fpath) if fname.lower().endswith(".csv")
                  else pd.read_excel(fpath))
            isin_col  = detect_isin_col(df)
            units_col = detect_units_col(df, isin_col) if isin_col else None
            if not isin_col or not units_col:
                print(f"    SKIP {fname}: could not detect ISIN/Units cols")
                continue
            for _, r in df.iterrows():
                isin  = str(r[isin_col]).strip()
                units = r[units_col]
                if ISIN_PAT.match(isin) and pd.notna(units) and float(units) > 0:
                    rows.append((client_id, isin, float(units)))
        except Exception as e:
            print(f"    ERROR {fname}: {e}")
    print(f"  Loaded {len(rows)} holdings across all clients")
    return rows

# ── Copy full latestNAV_Reports sheet into Check.xlsx ────────────────────────

def copy_nav_sheet(wb, master_path):
    """
    Copy latestNAV_Reports and Scheme types sheets from master into Check.xlsx.

    Reads formula strings (not data_only) so Final columns U-X are copied
    as live Excel formulae. Excel recalculates them on open.

    Scheme types is needed because col I in latestNAV_Reports references it
    via VLOOKUP, which the Final column formulas in U-X depend on.
    """
    # Read formula strings, not cached values
    master_wb = load_workbook(master_path, read_only=True)

    for sheet_name in ["Scheme types", NAV_SHEET]:
        if sheet_name not in master_wb.sheetnames:
            print(f"  WARNING: sheet {sheet_name!r} not found in master")
            continue
        master_ws = master_wb[sheet_name]
        ws = wb.create_sheet(sheet_name)
        for row in master_ws.iter_rows(values_only=True):
            ws.append(list(row))
        if sheet_name == NAV_SHEET:
            ws.freeze_panes = "E2"
        print(f"  {sheet_name}: copied {ws.max_row:,} rows from master")

    master_wb.close()




def write_all_client_holdings(wb, holdings_rows):
    """Write or overwrite AllClientHoldings sheet."""
    if HOLDINGS_SHEET in wb.sheetnames:
        del wb[HOLDINGS_SHEET]
    ws = wb.create_sheet(HOLDINGS_SHEET)

    # Headers
    for col, header in enumerate(["Key", "Client_ID", "ISIN", "Units"], 1):
        c = ws.cell(1, col)
        c.value = header
        apply(c, {**hdr_style(MID_BLUE), "font": Font(bold=True, color=WHITE, name="Arial", size=9)})
        ws.column_dimensions[c.column_letter].width = [22, 20, 16, 12][col-1]

    # Group by client to add row numbers
    from collections import defaultdict
    client_rows = defaultdict(list)
    for client_id, isin, units in holdings_rows:
        client_rows[client_id].append((isin, units))

    row = 2
    for client_id in sorted(client_rows.keys()):
        for seq, (isin, units) in enumerate(client_rows[client_id], 1):
            key = f"{client_id}|{seq}"
            for col, val in enumerate([key, client_id, isin, units], 1):
                c = ws.cell(row, col)
                c.value = val
                c.font  = Font(name="Arial", size=9)
                if col == 4:
                    c.number_format = "#,##0.0000"
                    c.alignment = Alignment(horizontal="right")
            row += 1

    # Freeze header row
    ws.freeze_panes = "A2"
    print(f"  AllClientHoldings: {row-2} rows written")
    return sorted(client_rows.keys())  # return list of client IDs

# ── Write Check Tab ──────────────────────────────────────────────────────────


def write_client_risk_map_sheet(wb, client_risk: dict[str, str]) -> None:
    """
    client_id → Averse|Moderate|Aggressive (one row per client, scales beyond
    hidden-column limits). Sheet used by Check Tab when Transactions is empty.
    """
    if CLIENT_RISK_MAP_SHEET in wb.sheetnames:
        del wb[CLIENT_RISK_MAP_SHEET]
    ws = wb.create_sheet(CLIENT_RISK_MAP_SHEET)
    ws["A1"].value = "client_id"
    ws["B1"].value = "risk_type"
    r = 2
    for cid in sorted(client_risk.keys(), key=lambda x: x.lower()):
        ws.cell(r, 1, value=cid)
        ws.cell(r, 2, value=client_risk[cid])
        r += 1
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 14
    ws.freeze_panes = "A2"


def risk_type_formula_from_transactions(
    tx_ws,
    *,
    transactions_sheet_name: str = TRANSACTIONS_SHEET_FOR_FORMULA,
    archetype_fallback: str = "Moderate",
) -> str | None:
    """
    VLOOKUP Client ID (C4) against Transactions: first row of each client matches
    that client's target_policy / risk_type (column target_policy after bulk_run).
    """
    if tx_ws is None or (tx_ws.max_row or 0) < 2:
        return None
    hdr: dict[str, int] = {}
    for c in range(1, (tx_ws.max_column or 0) + 1):
        v = tx_ws.cell(1, c).value
        if v is not None and str(v).strip():
            hdr[str(v).strip()] = c
    if "client_id" not in hdr or "target_policy" not in hdr:
        return None
    c_lo = hdr["client_id"]
    c_hi = hdr["target_policy"]
    if c_hi < c_lo:
        return None
    lo = get_column_letter(c_lo)
    hi = get_column_letter(c_hi)
    last = max(2, tx_ws.max_row)
    col_idx = c_hi - c_lo + 1
    esc = str(archetype_fallback or "Moderate").replace('"', '""')
    sh = transactions_sheet_name.replace("'", "''")
    return (
        f'=IFERROR(VLOOKUP($C$4,\'{sh}\'!${lo}$2:${hi}${last},{col_idx},FALSE),"{esc}")'
    )


def risk_type_formula_from_client_risk_map(
    last_data_row: int,
    archetype_fallback: str = "Moderate",
) -> str:
    esc = str(archetype_fallback or "Moderate").replace('"', '""')
    lr = max(2, last_data_row)
    sn = CLIENT_RISK_MAP_SHEET.replace("'", "''")
    return (
        f'=IFERROR(VLOOKUP($C$4,\'{sn}\'!$A$2:$B${lr},2,FALSE),"{esc}")'
    )


def write_check_tab(
    wb,
    client_ids,
    nav_sheet_name=NAV_SHEET,
    *,
    risk_type_cell_formula: str | None = None,
    archetype_fallback: str = "Moderate",
):
    """
    Build the full Check Tab with all formulas.

    If ``risk_type_cell_formula`` is set, cell C5 uses it (risk from Transactions or
    Client risk map sheet). Otherwise C5 is a manual archetype dropdown.
    """
    if CHECK_SHEET in wb.sheetnames:
        del wb[CHECK_SHEET]
    ws = wb.create_sheet(CHECK_SHEET, 0)  # first sheet

    # Column widths
    col_widths = {
        "A": 3,   # spacer
        "B": 24,  # labels
        "C": 22,  # inputs / values
        "D": 42,  # scheme name (wide)
        "E": 12,  # units
        "F": 12,  # NAV
        "G": 14,  # value
        "H": 10,  # eq%
        "I": 10,  # def%
        "J": 10,  # other%
        "K": 14,  # eq val
        "L": 14,  # def val
        "M": 14,  # other val
        "N": 10,  # weight%
    }
    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = width

    ws.row_dimensions[1].height = 24
    ws.row_dimensions[3].height = 18
    ws.row_dimensions[12].height = 18
    ws.freeze_panes = "B14"

    NAV = nav_sheet_name
    ACH = HOLDINGS_SHEET

    # ── Row 1: Title ──────────────────────────────────────────────────────────
    ws.merge_cells("B1:N1")
    c = ws["B1"]
    c.value = "MF Portfolio Rebalancer — Client Check Tab"
    c.font  = Font(bold=True, name="Arial", size=12, color=WHITE)
    c.fill  = PatternFill("solid", fgColor=DARK_BLUE)
    c.alignment = Alignment(horizontal="center", vertical="center")

    # ── Rows 3-11: Summary / Inputs ───────────────────────────────────────────
    first_client = client_ids[0] if client_ids else ""
    use_risk_formula = bool(risk_type_cell_formula and str(risk_type_cell_formula).strip())
    ws.merge_cells("B3:D3")
    c = ws["B3"]
    c.value = (
        "▶  INPUTS  (yellow = Client ID, New cash, etc.; blue = risk from Transactions / Client risk map)"
        if use_risk_formula
        else "▶  INPUTS  (yellow cells are editable)"
    )
    c.font = Font(bold=True, name="Arial", size=9, color=WHITE)
    c.fill = PatternFill("solid", fgColor=MID_BLUE)

    # Archetype lookup table (hidden, cols P-R)
    arch_data = [
        ("Averse",     0.35, 0.60, 0.05),
        ("Moderate",   0.50, 0.45, 0.05),
        ("Aggressive", 0.65, 0.30, 0.05),
    ]
    ws["P1"].value = "Archetype"
    ws["Q1"].value = "Target Eq"
    ws["R1"].value = "Target Def"
    ws["S1"].value = "Target Other"
    for i, (name, eq, df, oth) in enumerate(arch_data, 2):
        ws[f"P{i}"].value = name
        ws[f"Q{i}"].value = eq
        ws[f"R{i}"].value = df
        ws[f"S{i}"].value = oth
    # Hide the lookup columns
    ws.column_dimensions["P"].hidden = True
    ws.column_dimensions["Q"].hidden = True
    ws.column_dimensions["R"].hidden = True
    ws.column_dimensions["S"].hidden = True

    # Input rows
    row5_label = "Risk type (from run)" if use_risk_formula else "Archetype"
    inputs = [
        (4,  "Client ID",           first_client,   True),
        (5,  row5_label,              "Moderate",     True),
        (6,  "New Cash Addition (₹)", 0,            True),
        (7,  "Use Existing Schemes?","Yes",         True),
        (8,  "Run Date",            "=TODAY()",     False),
    ]
    for row, label, default, editable in inputs:
        c_lbl = ws.cell(row, 2)
        c_lbl.value = label
        apply(c_lbl, label_style())

        if row == 5 and use_risk_formula:
            c_val = ws.cell(row, 3)
            c_val.value = risk_type_cell_formula
            apply(c_val, formula_style(LIGHT_BLUE))
            continue

        c_val = ws.cell(row, 3)
        c_val.value = default
        if editable:
            apply(c_val, input_style())
        else:
            c_val.font = Font(name="Arial", size=9)
            if row == 8:
                c_val.number_format = "DD-MMM-YYYY"

    # Data validation: Client ID dropdown
    client_list_str = ",".join(client_ids[:50])  # Excel DV limit
    dv_client = DataValidation(
        type="list",
        formula1=f'"{client_list_str}"',
        showDropDown=False,
        showErrorMessage=True,
        errorTitle="Invalid",
        error="Select a client from the list",
    )
    ws.add_data_validation(dv_client)
    dv_client.add("C4")

    # Data validation: Archetype dropdown (manual mode only)
    if not use_risk_formula:
        dv_arch = DataValidation(
            type="list",
            formula1='"Averse,Moderate,Aggressive"',
            showDropDown=False,
        )
        ws.add_data_validation(dv_arch)
        dv_arch.add("C5")

    # ── Rows 9-11: Summary computed ───────────────────────────────────────────
    # Target (from archetype VLOOKUP into hidden table)
    summary_rows = [
        (9,  "Target Equity %",
              '=IFERROR(VLOOKUP($C$5,$P$2:$S$4,2,0),"")',
              '=IFERROR(VLOOKUP($C$5,$P$2:$S$4,3,0),"")',
              '=IFERROR(VLOOKUP($C$5,$P$2:$S$4,4,0),"")'),
        (10, "Current Portfolio",
              f'=IFERROR(SUMPRODUCT(($K${HOLDINGS_START}:$K${HOLDINGS_END})/MAX(1,SUMPRODUCT($G${HOLDINGS_START}:$G${HOLDINGS_END}))),"0.0%")',
              f'=IFERROR(SUMPRODUCT(($L${HOLDINGS_START}:$L${HOLDINGS_END})/MAX(1,SUMPRODUCT($G${HOLDINGS_START}:$G${HOLDINGS_END}))),"0.0%")',
              f'=IFERROR(SUMPRODUCT(($M${HOLDINGS_START}:$M${HOLDINGS_END})/MAX(1,SUMPRODUCT($G${HOLDINGS_START}:$G${HOLDINGS_END}))),"0.0%")'),
        (11, "Gap (Target - Current)",
              f'=IFERROR(C9-C10,"")',
              f'=IFERROR(D9-D10,"")',
              f'=IFERROR(E9-E10,"")'),
    ]

    # Sub-headers for columns C/D/E in summary
    for col, label in [(3,"Equity %"), (4,"Defensive %"), (5,"Others %")]:
        c = ws.cell(8, col)  # use row 8 col C/D/E as sub-headers if row 8 used
        # Actually put sub-headers in a label row
    # Put col headers at row 8 alongside Run Date
    for col, label in [(3,"Equity %"), (4,"Defensive %"), (5,"Others %")]:
        # Label row for the summary cols
        pass  # we'll label inline

    for row, label, eq_f, def_f, oth_f in summary_rows:
        ws.row_dimensions[row].height = 16
        c_lbl = ws.cell(row, 2)
        c_lbl.value = label
        apply(c_lbl, label_style())

        fill = LIGHT_BLUE if row == 9 else (GREEN if row == 10 else AMBER)
        for col, formula in [(3, eq_f), (4, def_f), (5, oth_f)]:
            c = ws.cell(row, col)
            c.value = formula
            c.number_format = "0.0%"
            apply(c, formula_style(fill))

    # Switch amount row
    ws.row_dimensions[11].height = 16
    total_val_formula = f'=IFERROR(SUM($G${HOLDINGS_START}:$G${HOLDINGS_END})+$C$6,0)'
    ws["B11"].value = "Switch Amount (₹)"
    apply(ws["B11"], label_style())
    ws["C11"].value = f'=IFERROR(ABS(C11_placeholder),0)'
    # Simpler: gap in INR
    ws["C11"].value = f'=IFERROR(ABS(C10-C9)*({total_val_formula}),0)'
    ws["C11"].number_format = "₹#,##0"
    apply(ws["C11"], formula_style(AMBER))

    # ── Row 12: Holdings section header ───────────────────────────────────────
    ws.merge_cells("B12:N12")
    c = ws["B12"]
    c.value = "▶  CLIENT HOLDINGS"
    c.font  = Font(bold=True, name="Arial", size=9, color=WHITE)
    c.fill  = PatternFill("solid", fgColor=MID_BLUE)

    # ── Row 13: Column headers ─────────────────────────────────────────────────
    col_headers = [
        (2, "#"),
        (3, "ISIN"),
        (4, "Scheme Name"),
        (5, "Units"),
        (6, "NAV (₹)"),
        (7, "Value (₹)"),
        (8, "Equity %"),
        (9, "Defensive %"),
        (10,"Others %"),
        (11,"Equity Val (₹)"),
        (12,"Defensive Val (₹)"),
        (13,"Others Val (₹)"),
        (14,"Weight %"),
    ]
    for col, header in col_headers:
        c = ws.cell(13, col)
        c.value = header
        apply(c, hdr_style(DARK_BLUE, size=8))

    # ── Rows 14-163: Holdings data rows ───────────────────────────────────────
    # Key formula: look up ClientID|RowNum in AllClientHoldings to get ISIN
    # Then VLOOKUP ISIN into latestNAV_Reports for scheme name, NAV, Eq%, Def%, Other%

    for i in range(1, MAX_HOLDINGS_ROWS + 1):
        row = HOLDINGS_START + i - 1

        # Row number
        ws.cell(row, 2).value = i
        ws.cell(row, 2).font  = Font(name="Arial", size=8)
        ws.cell(row, 2).alignment = Alignment(horizontal="center")

        # Col C: ISIN via VLOOKUP on key (ClientID|RowNum) into AllClientHoldings
        key_formula = f'$C$4&"|"&{i}'
        isin_formula = (
            f'=IFERROR(VLOOKUP({key_formula},{ACH}!$A:$D,3,0),"")'
        )
        c_isin = ws.cell(row, 3)
        c_isin.value = isin_formula
        c_isin.font  = Font(name="Arial", size=8, color="000080")
        c_isin.alignment = Alignment(horizontal="left")

        # Col D: Scheme Name — col D(4) is left of col E(5), need INDEX/MATCH
        name_formula = (
            f'=IFERROR(IF(C{row}="","",INDEX({nav_sheet_name}!$D:$D,'
            f'MATCH(C{row},{nav_sheet_name}!$E:$E,0))),"")'
        )
        c_name = ws.cell(row, 4)
        c_name.value = name_formula
        c_name.font  = Font(name="Arial", size=8)
        c_name.alignment = Alignment(horizontal="left")

        # Col E: Units — from AllClientHoldings col D
        units_formula = (
            f'=IFERROR(IF(C{row}="","",VLOOKUP({key_formula},{ACH}!$A:$D,4,0)),"")'
        )
        c_units = ws.cell(row, 5)
        c_units.value = units_formula
        c_units.number_format = "#,##0.0000"
        c_units.font  = Font(name="Arial", size=8)
        c_units.alignment = Alignment(horizontal="right")

        # Col F: NAV — col G(7), offset from E(5) = 3
        nav_formula = (
            f'=IFERROR(IF(C{row}="","",VLOOKUP(C{row},{nav_sheet_name}!$E:$G,3,0)),"")'
        )
        c_nav = ws.cell(row, 6)
        c_nav.value = nav_formula
        c_nav.number_format = "#,##0.0000"
        c_nav.font  = Font(name="Arial", size=8)
        c_nav.alignment = Alignment(horizontal="right")

        # Col G: Value = Units * NAV
        val_formula = f'=IFERROR(IF(C{row}="","",E{row}*F{row}),"")'
        c_val = ws.cell(row, 7)
        c_val.value = val_formula
        c_val.number_format = "#,##0"
        c_val.font  = Font(name="Arial", size=8)
        c_val.alignment = Alignment(horizontal="right")

        # Col H: Final Equity % — col U(21), offset from E(5) = 17
        eq_formula = (
            f'=IFERROR(IF(C{row}="","",VLOOKUP(C{row},{nav_sheet_name}!$E:$U,17,0)),"")'
        )
        c_eq = ws.cell(row, 8)
        c_eq.value = eq_formula
        c_eq.number_format = "0.0%"
        c_eq.font  = Font(name="Arial", size=8)
        c_eq.fill  = PatternFill("solid", fgColor="F0FFF0")
        c_eq.alignment = Alignment(horizontal="right")

        # Col I: Defensive % = Final Debt%(V=22,off=18) + Final Cash%(W=23,off=19)
        def_formula = (
            f'=IFERROR(IF(C{row}="","",VLOOKUP(C{row},{nav_sheet_name}!$E:$V,18,0)'
            f'+VLOOKUP(C{row},{nav_sheet_name}!$E:$W,19,0)),"")'
        )
        c_def = ws.cell(row, 9)
        c_def.value = def_formula
        c_def.number_format = "0.0%"
        c_def.font  = Font(name="Arial", size=8)
        c_def.fill  = PatternFill("solid", fgColor="F0F0FF")
        c_def.alignment = Alignment(horizontal="right")

        # Col J: Others % — col X(24), offset from E(5) = 20
        oth_formula = (
            f'=IFERROR(IF(C{row}="","",VLOOKUP(C{row},{nav_sheet_name}!$E:$X,20,0)),"")'
        )
        c_oth = ws.cell(row, 10)
        c_oth.value = oth_formula
        c_oth.number_format = "0.0%"
        c_oth.font  = Font(name="Arial", size=8)
        c_oth.fill  = PatternFill("solid", fgColor="FFFFF0")
        c_oth.alignment = Alignment(horizontal="right")

        # Col K: Equity Value = Value * Equity%
        eq_val_formula = f'=IFERROR(IF(C{row}="","",G{row}*H{row}),"")'
        c_eqval = ws.cell(row, 11)
        c_eqval.value = eq_val_formula
        c_eqval.number_format = "#,##0"
        c_eqval.font  = Font(name="Arial", size=8)
        c_eqval.alignment = Alignment(horizontal="right")

        # Col L: Defensive Value = Value * Defensive%
        def_val_formula = f'=IFERROR(IF(C{row}="","",G{row}*I{row}),"")'
        c_defval = ws.cell(row, 12)
        c_defval.value = def_val_formula
        c_defval.number_format = "#,##0"
        c_defval.font  = Font(name="Arial", size=8)
        c_defval.alignment = Alignment(horizontal="right")

        # Col M: Others Value = Value * Others%
        oth_val_formula = f'=IFERROR(IF(C{row}="","",G{row}*J{row}),"")'
        c_othval = ws.cell(row, 13)
        c_othval.value = oth_val_formula
        c_othval.number_format = "#,##0"
        c_othval.font  = Font(name="Arial", size=8)
        c_othval.alignment = Alignment(horizontal="right")

        # Col N: Weight % = Value / Total Portfolio Value
        total_val = f'SUM($G${HOLDINGS_START}:$G${HOLDINGS_END})'
        wt_formula = f'=IFERROR(IF(C{row}="","",G{row}/({total_val})),"")'
        c_wt = ws.cell(row, 14)
        c_wt.value = wt_formula
        c_wt.number_format = "0.0%"
        c_wt.font  = Font(name="Arial", size=8)
        c_wt.alignment = Alignment(horizontal="right")

        # Alternate row shading
        if i % 2 == 0:
            for col in range(2, 15):
                cell = ws.cell(row, col)
                if not cell.fill or cell.fill.fgColor.rgb == "00000000":
                    cell.fill = PatternFill("solid", fgColor=LIGHT_GREY)

    # ── Row HOLDINGS_END+1: Totals row ────────────────────────────────────────
    total_row = HOLDINGS_END + 1
    ws.row_dimensions[total_row].height = 16
    c_tot_lbl = ws.cell(total_row, 3)
    c_tot_lbl.value = "TOTAL"
    c_tot_lbl.font  = Font(bold=True, name="Arial", size=9)

    for col, col_letter in [(7,"G"), (11,"K"), (12,"L"), (13,"M")]:
        c = ws.cell(total_row, col)
        c.value = f'=SUM({col_letter}{HOLDINGS_START}:{col_letter}{HOLDINGS_END})'
        c.number_format = "#,##0"
        c.font  = Font(bold=True, name="Arial", size=9)
        c.fill  = PatternFill("solid", fgColor=DARK_BLUE)
        c.font  = Font(bold=True, name="Arial", size=9, color=WHITE)
        c.alignment = Alignment(horizontal="right")

    # ── SUMPRODUCT for current portfolio % (used in summary row 10) ──────────
    # Fix the summary row 10 formulas now that we know the column letters
    total_G = f'SUM($G${HOLDINGS_START}:$G${HOLDINGS_END})'
    ws["C10"].value = f'=IFERROR(SUMPRODUCT($G${HOLDINGS_START}:$G${HOLDINGS_END},$H${HOLDINGS_START}:$H${HOLDINGS_END})/({total_G}),0)'
    ws["D10"].value = f'=IFERROR(SUMPRODUCT($G${HOLDINGS_START}:$G${HOLDINGS_END},$I${HOLDINGS_START}:$I${HOLDINGS_END})/({total_G}),0)'
    ws["E10"].value = f'=IFERROR(SUMPRODUCT($G${HOLDINGS_START}:$G${HOLDINGS_END},$J${HOLDINGS_START}:$J${HOLDINGS_END})/({total_G}),0)'
    for cell_ref in ["C10","D10","E10"]:
        ws[cell_ref].number_format = "0.0%"
        apply(ws[cell_ref], formula_style(GREEN))

    # Fix gap row (row 11)
    ws["C11"].value = f'=IFERROR(C9-C10,"")'
    ws["D11"].value = f'=IFERROR(D9-D10,"")'
    ws["E11"].value = f'=IFERROR(E9-E10,"")'
    for cell_ref in ["C11","D11","E11"]:
        ws[cell_ref].number_format = "0.0%"
        apply(ws[cell_ref], formula_style(AMBER))

    # Switch amount
    ws["C11_switch"] if False else None  # placeholder
    switch_row = 11
    # Add switch amount to col G row 9-11 area
    ws["G9"].value  = "Portfolio Value (₹)"
    ws["G9"].font   = Font(bold=True, name="Arial", size=8)
    ws["H9"].value  = f'=IFERROR({total_G}+$C$6,0)'
    ws["H9"].number_format = "₹#,##0"
    ws["H9"].font   = Font(name="Arial", size=9)

    ws["G10"].value = "Switch Amount (₹)"
    ws["G10"].font  = Font(bold=True, name="Arial", size=8)
    ws["H10"].value = (
        f'=IFERROR(ABS(C9-C10)*H9,0)'
    )
    ws["H10"].number_format = "₹#,##0"
    ws["H10"].font  = Font(name="Arial", size=9)
    ws["H10"].fill  = PatternFill("solid", fgColor=AMBER)

    # ── Transactions section ──────────────────────────────────────────────────
    ws.merge_cells(f"B{TXNS_START}:N{TXNS_START}")
    c = ws[f"B{TXNS_START}"]
    c.value = "▶  SUGGESTED TRANSACTIONS  (written by bulk_run.py)"
    c.font  = Font(bold=True, name="Arial", size=9, color=WHITE)
    c.fill  = PatternFill("solid", fgColor=MID_BLUE)

    txn_headers = ["#","ISIN","Scheme Name","Action","Amount (₹)"]
    for col, header in enumerate(txn_headers, 2):
        c = ws.cell(TXNS_START+1, col)
        c.value = header
        apply(c, hdr_style(DARK_BLUE, size=8))

    # 20 blank transaction rows
    for i in range(1, 21):
        r = TXNS_START + 1 + i
        ws.cell(r, 2).value = i
        ws.cell(r, 2).font  = Font(name="Arial", size=8)

    print(f"  Check Tab: {MAX_HOLDINGS_ROWS} holdings rows + transactions section written")

# ── Main ─────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(
        description="Build Check Tab Excel workbook from client holdings."
    )
    parser.add_argument(
        "--master",
        default=None,
        help="NAV master workbook (default: data/latestNAV_Reports.xlsx, else newest backup)",
    )
    parser.add_argument("--clients", default=DEFAULT_CLIENTS,
                        help=f"Client holdings directory (default: {DEFAULT_CLIENTS})")
    parser.add_argument("--output",  default=DEFAULT_OUTPUT,
                        help=f"Output Check.xlsx path (default: {DEFAULT_OUTPUT})")
    args = parser.parse_args()
    if args.master is None:
        args.master = default_master()

    if not os.path.exists(args.master):
        print(f"ERROR: Master file not found: {args.master}")
        print("Available backup files in data/:")
        for f in sorted(os.listdir("data")):
            if f.startswith("latestNAV") and f.endswith(".xlsx"):
                print(f"  data/{f}")
        sys.exit(1)

    if not os.path.isdir(args.clients):
        print(f"ERROR: Client directory not found: {args.clients}")
        sys.exit(1)

    print(f"\n=== Build Check Tab ===")
    print(f"  NAV master (reference): {args.master}")
    print(f"  Client holdings dir:    {args.clients}")
    print(f"  Output:                 {args.output}")

    # Load client data
    holdings_rows = load_all_clients(args.clients)
    if not holdings_rows:
        print("ERROR: No client holdings found")
        sys.exit(1)

    master_read = args.master
    temp_copy = None
    if _same_nav_path(args.master, LIVE_NAV_REPORTS):
        fd, temp_copy = tempfile.mkstemp(suffix=".xlsx")
        os.close(fd)
        shutil.copyfile(args.master, temp_copy)
        master_read = temp_copy
        print(
            "  NOTE: Master path is live latestNAV_Reports.xlsx — "
            "reading via temporary copy."
        )

    # Create a brand new workbook — do NOT touch the master file
    from openpyxl import Workbook
    wb = Workbook()
    wb.remove(wb.active)
    print(f"  Creating new workbook: {args.output}")

    # Copy the full latestNAV_Reports sheet into Check.xlsx
    # This keeps all original column positions intact so VLOOKUP offsets
    # are identical — no remapping, fully inspectable in Excel
    try:
        print(f"  Copying {NAV_SHEET} sheet from master...")
        copy_nav_sheet(wb, master_read)
    finally:
        if temp_copy and os.path.exists(temp_copy):
            try:
                os.unlink(temp_copy)
            except OSError:
                pass

    # Write AllClientHoldings
    print(f"  Writing {HOLDINGS_SHEET}...")
    client_ids = write_all_client_holdings(wb, holdings_rows)
    print(f"  Client IDs: {client_ids}")

    # Write Check Tab — formulas reference latestNAV_Reports sheet by name
    print(f"  Writing {CHECK_SHEET}...")
    write_check_tab(wb, client_ids, nav_sheet_name=NAV_SHEET)

    # Save to output path (not the master)
    os.makedirs(os.path.dirname(args.output) if os.path.dirname(args.output) else ".", exist_ok=True)
    wb.save(args.output)
    print(f"\n  ✓ Saved: {args.output}")
    print(f"  latestNAV_Reports.xlsx was NOT modified.")
    print(f"  Open {args.output} in Excel to use the Check Tab.")
    print(f"  Select a client from the dropdown in C4 to populate holdings.")
    print(f"\n=== Done ===\n")

if __name__ == "__main__":
    main()