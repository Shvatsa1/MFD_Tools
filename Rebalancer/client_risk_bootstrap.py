"""
Build or refresh client_risk_pref.xlsx: sheet ``First steps`` (onboarding), ``ClientAges``
(data + risk_preference dropdowns).

Each file stem in the holdings folder becomes a client_id row. Optional client_name is sniffed.
age and risk_preference merge from an existing workbook when ``merge=True``.
"""

from __future__ import annotations

import os

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font as XLFont
from openpyxl.worksheet.datavalidation import DataValidation

from bulk_run import CLIENT_AGES_SHEET, ISIN_PATTERN

RISK_PREF_LIST = "Averse,Moderate,Aggressive"
FIRST_STEPS_SHEET = "First steps"


def _finalize_client_risk_workbook(
    path: str,
    *,
    n_client_rows: int,
) -> None:
    """
    Insert ``First steps`` (first-time instructions), list validation on column D
    (risk_preference), freeze header row on ClientAges.
    """
    wb = load_workbook(path)
    if FIRST_STEPS_SHEET in wb.sheetnames:
        del wb[FIRST_STEPS_SHEET]
    inst = wb.create_sheet(FIRST_STEPS_SHEET, 0)
    inst.column_dimensions["A"].width = 96
    t = inst.cell(1, 1)
    t.value = "Client risk file — first-time setup"
    t.font = XLFont(bold=True, name="Arial", size=12)
    lines = [
        "",
        "• Column D (risk_preference): use the dropdown in each row — Averse, Moderate, or Aggressive.",
        "• Column C (age): enter a number for every client when you enable age_based in mfd_pack.ini.",
        "• Sheet ClientAges: one row per holdings file (client id = filename stem). Re-run bootstrap after adding clients.",
        "",
        "Next: save this workbook. In mfd_pack.ini set client_risk_pref (or client_ages) to this file path;",
        "uncomment age_based if you use glide (100 − k × age). Run the rebalancer. The pack Check Tab will",
        "show Risk type (from pref) for the Client ID you pick, synced from this file.",
    ]
    for i, line in enumerate(lines, start=2):
        c = inst.cell(i, 1)
        c.value = line
        c.font = XLFont(name="Arial", size=10)
        c.alignment = Alignment(wrap_text=True, vertical="top")

    ws = wb[CLIENT_AGES_SHEET]
    ws.freeze_panes = "A2"
    first_row = 2
    last_row = max(first_row, 1 + max(n_client_rows, 1), 500)
    dv = DataValidation(
        type="list",
        formula1=f'"{RISK_PREF_LIST}"',
        allow_blank=True,
        showErrorMessage=True,
        errorTitle="Risk preference",
        error="Choose Averse, Moderate, or Aggressive.",
    )
    ws.add_data_validation(dv)
    dv.add(f"D{first_row}:D{last_row}")
    wb.save(path)
    wb.close()

SKIP_STEMS = frozenset(
    {
        "client_ages",
        "client_risk_pref",
        "fund_choices",
    }
)


def _iter_holdings_files(client_dir: str) -> list[tuple[str, str]]:
    out: list[tuple[str, str]] = []
    for fname in sorted(os.listdir(client_dir)):
        if fname.startswith("~$"):
            continue
        lower = fname.lower()
        if not lower.endswith((".xlsx", ".xls", ".csv")):
            continue
        stem = os.path.splitext(fname)[0]
        if stem.lower() in SKIP_STEMS:
            continue
        out.append((stem, os.path.join(client_dir, fname)))
    return out


def sniff_client_name(fpath: str) -> str:
    try:
        if fpath.lower().endswith(".csv"):
            df = pd.read_csv(fpath, nrows=40)
        else:
            df = pd.read_excel(fpath, nrows=40)
    except Exception:
        return ""
    if df.empty:
        return ""
    norm = {str(c).strip().lower(): c for c in df.columns}
    for key in (
        "client name",
        "investor name",
        "name",
        "investor",
        "client",
    ):
        if key not in norm:
            continue
        col = norm[key]
        for v in df[col]:
            if pd.isna(v):
                continue
            s = str(v).strip()
            if not s or s.lower() == "nan":
                continue
            if ISIN_PATTERN.match(s):
                continue
            return s
    return ""


def _load_existing_sidecar(path: str) -> dict[str, dict]:
    if not os.path.isfile(path):
        return {}
    try:
        df = pd.read_excel(path, sheet_name=CLIENT_AGES_SHEET)
    except Exception:
        return {}
    norm = {str(c).strip().lower().replace(" ", "_"): c for c in df.columns}

    def pick(*names):
        for n in names:
            if n in norm:
                return norm[n]
        return None

    c_id = pick("client_id", "clientid")
    if not c_id:
        return {}
    age_c = pick("age")
    risk_c = pick("risk_preference", "riskpreference", "preference")
    name_c = pick("client_name", "clientname", "name")
    out: dict[str, dict] = {}
    for _, row in df.iterrows():
        cid = str(row[c_id]).strip()
        if not cid or cid.lower() == "nan":
            continue
        item: dict = {}
        if age_c is not None and age_c in row.index:
            item["age"] = row[age_c]
        if risk_c is not None and risk_c in row.index:
            v = row[risk_c]
            item["risk_preference"] = (
                "" if pd.isna(v) else str(v).strip()
            )
        if name_c is not None and name_c in row.index:
            v = row[name_c]
            item["client_name"] = (
                "" if pd.isna(v) else str(v).strip()
            )
        out[cid] = item
    return out


def build_client_risk_pref_workbook(
    client_dir: str,
    output_path: str,
    *,
    merge: bool = True,
) -> int:
    """
    Write workbook with sheet CLIENT_AGES_SHEET and columns:
    client_id, client_name, age, risk_preference.

    Returns number of client rows written.
    """
    files = _iter_holdings_files(client_dir)
    if not files:
        raise ValueError(f"No client holdings files found in {client_dir!r}")

    existing: dict[str, dict] = {}
    if merge:
        existing = _load_existing_sidecar(output_path)

    rows = []
    for client_id, fpath in files:
        sniffed = sniff_client_name(fpath)
        prev = existing.get(client_id, {})
        name = str(prev.get("client_name", "") or "").strip() or sniffed
        age = prev.get("age", "")
        risk = str(prev.get("risk_preference", "") or "").strip()
        rows.append(
            {
                "client_id": client_id,
                "client_name": name,
                "age": age,
                "risk_preference": risk,
            }
        )

    rows.sort(key=lambda r: r["client_id"].lower())
    df = pd.DataFrame(rows)
    os.makedirs(os.path.dirname(os.path.abspath(output_path)) or ".", exist_ok=True)
    with pd.ExcelWriter(
        output_path, engine="openpyxl", mode="w"
    ) as writer:
        df.to_excel(writer, sheet_name=CLIENT_AGES_SHEET, index=False)
    n = len(rows)
    _finalize_client_risk_workbook(output_path, n_client_rows=n)
    return n


def write_headers_only(path: str) -> None:
    """Empty template (headers only) for shipping inside a release pack."""
    df = pd.DataFrame(
        columns=["client_id", "client_name", "age", "risk_preference"]
    )
    os.makedirs(os.path.dirname(os.path.abspath(path)) or ".", exist_ok=True)
    with pd.ExcelWriter(path, engine="openpyxl", mode="w") as writer:
        df.to_excel(writer, sheet_name=CLIENT_AGES_SHEET, index=False)
    _finalize_client_risk_workbook(path, n_client_rows=0)
