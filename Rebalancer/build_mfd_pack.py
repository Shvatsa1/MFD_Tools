"""
build_mfd_pack.py
-----------------
Build a distributor-facing Excel pack + audit manifest in one step.

  • Format B only — one holdings file (xlsx/csv) per client in a folder.
  • Global archetype + global new-cash (same as bulk_run.py).
  • Transactions default to existing-holdings BUY pools only; optional --allow-new-funds
    + FundChoices workbook (see fund_choices.py).

Outputs (default paths can be overridden):
  1) generate_transactions.xlsx — Excel 2019–friendly workbook:
       Instructions, Parameters, Transactions,
       Scheme types, latestNAV_Reports, AllClientHoldings, Check Tab
  2) generate_transactions_manifest.json — run metadata for audit / rerun.

The Check Tab and NAV copy reuse build_check_tab.py (VLOOKUP / INDEX-MATCH only).

Requires: same preconditions as bulk_run (master with Final cols U–X from
add_final_columns.py), client files detectable by bulk_run column heuristics.

Usage:
  python build_mfd_pack.py --clients data/dummy_clients/by_client/
  python build_mfd_pack.py --clients path/to/clients --master data/latestNAV_Reports.xlsx \\
      --archetype Moderate --new-cash 0 --output output/mfd_pack.xlsx

  Optional: copy mfd_pack.example.ini to mfd_pack.ini next to the exe (or pass --config).
  INI holds clients_folder, master, output, archetype, new_cash, etc. Edit in Notepad.

  Step 0 (glide path): scan holdings folder and write client risk workbook, then exit:
  python build_mfd_pack.py --bootstrap-client-risk
  (uses clients_folder from INI or --clients; output from client_risk_pref / client_ages in INI,
   --bootstrap-output, or data/client_risk_pref.xlsx)

"""

from __future__ import annotations

import argparse
import configparser
import datetime
import json
import os
import shutil
import sys
import tempfile

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

import config as app_config
from bulk_run import bulk_run, load_client_ages
from build_check_tab import (
    CHECK_SHEET,
    CLIENT_RISK_MAP_SHEET,
    HOLDINGS_SHEET,
    LIVE_NAV_REPORTS,
    NAV_SHEET,
    _same_nav_path,
    copy_nav_sheet,
    default_master,
    load_all_clients,
    risk_type_formula_from_client_risk_map,
    risk_type_formula_from_transactions,
    write_all_client_holdings,
    write_check_tab,
    write_client_risk_map_sheet,
)
from client_risk_bootstrap import build_client_risk_pref_workbook
from portfolio import load_master_for_portfolio

INSTRUCTIONS_SHEET = "Instructions"
PARAMETERS_SHEET = "Parameters"
TRANSACTIONS_SHEET = "Transactions"

DEFAULT_OUTPUT_XLSX = os.path.join("output", "generate_transactions.xlsx")
DEFAULT_CLIENTS = os.path.join("data", "dummy_clients", "by_client")
PACK_INI_BASENAME = "mfd_pack.ini"


def _client_risk_for_check_tab(
    client_ids: list[str],
    client_ages_path: str | None,
    archetype_fallback: str,
) -> dict[str, str] | None:
    """
    client_id -> Averse|Moderate|Aggressive for Check Tab VLOOKUP.

    Used whenever ``client_ages_path`` points to a readable sidecar (not only when
    age_based glide is on), so the Check Tab matches client_risk_pref.xlsx.
    """
    if not client_ages_path or not os.path.isfile(client_ages_path):
        return None
    try:
        ages = load_client_ages(client_ages_path)
    except Exception:
        return None

    def norm(pref: object) -> str:
        t = str(pref).strip().lower()
        for k in app_config.ARCHETYPES:
            if k.lower() == t:
                return k
        return archetype_fallback

    out: dict[str, str] = {}
    for cid in client_ids:
        row = ages.get(cid)
        if row:
            out[cid] = norm(row.get("risk_preference", ""))
        else:
            out[cid] = archetype_fallback
    return out


def _extract_config_path(argv: list[str]) -> str | None:
    i = 0
    while i < len(argv):
        a = argv[i]
        if a == "--config" and i + 1 < len(argv):
            return argv[i + 1]
        if a.startswith("--config="):
            return a.split("=", 1)[1]
        i += 1
    return None


def resolve_mfd_pack_ini_path(explicit: str | None) -> str | None:
    """Return path to INI if it exists, else None."""
    if explicit:
        p = os.path.normpath(
            explicit if os.path.isabs(explicit) else os.path.join(os.getcwd(), explicit)
        )
        return p if os.path.isfile(p) else None
    cand = os.path.join(os.getcwd(), PACK_INI_BASENAME)
    return cand if os.path.isfile(cand) else None


def load_mfd_pack_ini(ini_path: str | None) -> dict:
    """
    Read [mfd_pack] section; return kwargs for argparse.set_defaults().
    Keys: clients, master, output, manifest, archetype, new_cash, defensive_buckets,
    age_based, client_ages, client_risk_pref (alias for client_ages), allow_new_funds, fund_choices,
    invest_in_existing_funds_only (inverts allow_new_funds).
    """
    if not ini_path:
        return {}
    cfg = configparser.ConfigParser(interpolation=None)
    try:
        # utf-8-sig strips a leading BOM (\ufeff) from Notepad / PowerShell Set-Content utf8
        data = cfg.read(ini_path, encoding="utf-8-sig")
        if not data:
            print(f"WARNING: Could not read INI: {ini_path}", file=sys.stderr)
            return {}
    except configparser.Error as e:
        print(f"WARNING: Invalid INI {ini_path}: {e}", file=sys.stderr)
        return {}
    if "mfd_pack" not in cfg:
        print(f"WARNING: INI missing [mfd_pack] section: {ini_path}", file=sys.stderr)
        return {}
    sec = cfg["mfd_pack"]

    def raw(key: str) -> str:
        return str(sec.get(key, "") or "").strip()

    out: dict = {}
    v = raw("clients_folder")
    if v:
        out["clients"] = v
    v = raw("master")
    if v:
        out["master"] = v
    v = raw("output")
    if v:
        out["output"] = v
    v = raw("manifest")
    if v:
        out["manifest"] = v
    v = raw("archetype")
    if v:
        out["archetype"] = v
    nc = raw("new_cash")
    if nc != "":
        try:
            out["new_cash"] = float(nc.replace(",", ""))
        except ValueError:
            print(f"WARNING: Invalid new_cash in INI: {nc!r}", file=sys.stderr)
    v = raw("defensive_buckets")
    if v:
        if v in ("debt_cash", "debt_cash_other"):
            out["defensive_buckets"] = v
        else:
            print(f"WARNING: Invalid defensive_buckets (use debt_cash or debt_cash_other): {v!r}", file=sys.stderr)
    ab = raw("age_based")
    if ab.lower() in ("1", "true", "yes", "on"):
        out["age_based"] = True
    elif ab.lower() in ("0", "false", "no", "off"):
        out["age_based"] = False
    crp = raw("client_risk_pref")
    ca = raw("client_ages")
    if ca:
        out["client_ages"] = ca
    elif crp:
        out["client_ages"] = crp
    ie = raw("invest_in_existing_funds_only")
    if ie:
        if ie.lower() in ("0", "false", "no", "off"):
            out["allow_new_funds"] = True
        elif ie.lower() in ("1", "true", "yes", "on"):
            out["allow_new_funds"] = False
    anf = raw("allow_new_funds")
    if anf:
        if anf.lower() in ("1", "true", "yes", "on"):
            out["allow_new_funds"] = True
        elif anf.lower() in ("0", "false", "no", "off"):
            out["allow_new_funds"] = False
    v = raw("fund_choices")
    if v:
        out["fund_choices"] = v
    return out


HEADER_FILL = "1F4E79"
HEADER_FONT = Font(bold=True, color="FFFFFF", name="Arial", size=10)


def _reorder_sheets(wb: Workbook, names_in_order: list[str]) -> None:
    """Set workbook sheet order; skip missing names (e.g. optional Scheme types)."""
    wb._sheets = [wb[n] for n in names_in_order if n in wb.sheetnames]


def _write_instructions(ws) -> None:
    ws.column_dimensions["A"].width = 100
    title = ws.cell(1, 1)
    title.value = "MF rebalancer — MFD pack (read-only review workbook)"
    title.font = Font(bold=True, name="Arial", size=14)

    lines = [
        "",
        "This file was generated by Python. You do not need to run code to review it.",
        "",
        "What to do:",
        "  1) Read Parameters for the exact inputs used for this run.",
        "  2) Open Transactions — filter by client_id to see suggested switches (₹).",
        "  3) Use Check Tab to inspect one client: pick Client ID, archetype, and new cash (what-if).",
        "     Note: Check Tab targets may differ from Parameters unless you match archetype / cash.",
        "",
        "Scope (version 1):",
        "  • Client holdings: one file per customer in the folder listed on Parameters.",
        "  • Buys / sells only across schemes the client already holds (no new fund names).",
        "  • One archetype applied to every client for this run (see Parameters).",
        "",
        "Master NAV data: sheets Scheme types + latestNAV_Reports are a frozen copy.",
        "Final classification uses columns U–X when present (from add_final_columns.py on the master).",
        "",
        "Regenerated packs: keep the matching *_manifest.json to know which master and folder were used.",
    ]
    for i, text in enumerate(lines, start=2):
        c = ws.cell(i, 1)
        c.value = text
        c.font = Font(name="Arial", size=10)
        c.alignment = Alignment(wrap_text=True, vertical="top")


def _write_parameters(ws, rows: list[tuple[str, str]]) -> None:
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 72
    ws["A1"].value = "Parameter"
    ws["B1"].value = "Value"
    for c in (ws["A1"], ws["B1"]):
        c.font = HEADER_FONT
        c.fill = PatternFill("solid", fgColor=HEADER_FILL)
    for i, (k, v) in enumerate(rows, start=2):
        ws.cell(i, 1).value = k
        ws.cell(i, 2).value = v
        ws.cell(i, 1).font = Font(bold=True, name="Arial", size=10)
        ws.cell(i, 2).font = Font(name="Arial", size=10)
        ws.cell(i, 2).alignment = Alignment(wrap_text=True, vertical="top")


def _write_transactions(wb: Workbook, df: pd.DataFrame) -> None:
    ws = wb[TRANSACTIONS_SHEET]
    if ws.max_row and ws.max_row > 0:
        ws.delete_rows(1, ws.max_row)
    if df is not None and not df.empty:
        for row in dataframe_to_rows(df, index=False, header=True):
            ws.append(row)
    else:
        empty_cols = [
            "client_id",
            "isin",
            "scheme_name",
            "action",
            "amount_inr",
            "target_policy",
            "current_equity_pct",
            "target_equity_pct",
            "current_defensive_pct",
            "target_defensive_pct",
            "current_other_pct",
            "target_other_pct",
            "portfolio_value_inr",
            "switch_amount_inr",
            "new_cash_inr",
            "deployed_new_cash_to_others_inr",
            "total_buy_inr",
            "total_sell_inr",
            "net_flow_inr",
        ]
        ws.append(empty_cols)
    # Header style
    for col in range(1, ws.max_column + 1):
        c = ws.cell(1, col)
        c.font = HEADER_FONT
        c.fill = PatternFill("solid", fgColor=HEADER_FILL)
    ws.freeze_panes = "A2"
    for col in range(1, ws.max_column + 1):
        ws.column_dimensions[get_column_letter(col)].width = min(18, max(12, len(str(ws.cell(1, col).value or "")) + 2))


def main() -> None:
    # PyInstaller one-file: cwd may be System32 if launched from Explorer;
    # default relative paths (config, data/, output/) must be next to the .exe.
    if getattr(sys, "frozen", False) and getattr(sys, "executable", None):
        os.chdir(os.path.dirname(os.path.abspath(sys.executable)))
    if sys.platform == "win32" and hasattr(sys.stdout, "reconfigure"):
        try:
            sys.stdout.reconfigure(encoding="utf-8", errors="replace")
            sys.stderr.reconfigure(encoding="utf-8", errors="replace")
        except (OSError, ValueError):
            pass

    cfg_arg = _extract_config_path(sys.argv[1:])
    ini_resolved = resolve_mfd_pack_ini_path(cfg_arg)
    if cfg_arg and ini_resolved is None:
        print(f"ERROR: --config file not found: {cfg_arg}", file=sys.stderr)
        sys.exit(1)
    ini_defaults = load_mfd_pack_ini(ini_resolved)

    parser = argparse.ArgumentParser(
        description="Build MFD Excel pack (Format B) + manifest.json"
    )
    parser.add_argument(
        "--config",
        default=None,
        help=f"Path to INI settings file (default: try ./{PACK_INI_BASENAME} if it exists)",
    )
    parser.add_argument(
        "--clients",
        default=DEFAULT_CLIENTS,
        help=f"Folder of client files (default: {DEFAULT_CLIENTS})",
    )
    parser.add_argument(
        "--master",
        default=None,
        help="latestNAV_Reports workbook (default: data/latestNAV_Reports.xlsx)",
    )
    parser.add_argument(
        "--output",
        default=DEFAULT_OUTPUT_XLSX,
        help=f"Output .xlsx path (default: {DEFAULT_OUTPUT_XLSX})",
    )
    parser.add_argument(
        "--manifest",
        default=None,
        help="Manifest JSON path (default: next to --output with _manifest.json suffix)",
    )
    parser.add_argument(
        "--archetype",
        default="Moderate",
        choices=list(app_config.ARCHETYPES.keys()),
    )
    parser.add_argument(
        "--new-cash",
        type=float,
        default=0,
        help="Fresh cash per client (INR), same as bulk_run",
    )
    parser.add_argument(
        "--defensive-buckets",
        choices=["debt_cash", "debt_cash_other"],
        default=None,
        help="Override config.DEFENSIVE_BUCKETS for this run (default: use config.py value)",
    )
    parser.add_argument(
        "--age-based",
        action="store_true",
        default=False,
        help="Per-client glide path (100−k×age) using client_ages workbook",
    )
    parser.add_argument(
        "--client-ages",
        default=None,
        dest="client_ages",
        help=f"Sidecar .xlsx with sheet ClientAges (default: {app_config.CLIENT_AGES_FILE})",
    )
    parser.add_argument(
        "--allow-new-funds",
        action="store_true",
        default=False,
        help="BUYs may use FundChoices sheet (see --fund-choices)",
    )
    parser.add_argument(
        "--fund-choices",
        default=None,
        dest="fund_choices",
        help=f"Workbook with FundChoices (default: {app_config.FUND_CHOICES_FILE})",
    )
    parser.add_argument(
        "--bootstrap-client-risk",
        action="store_true",
        help="Scan --clients folder and write client risk workbook (sheet ClientAges); exit.",
    )
    parser.add_argument(
        "--bootstrap-output",
        default=None,
        help="Output .xlsx for --bootstrap-client-risk (default: INI client_ages or client_risk_pref, else data/client_risk_pref.xlsx).",
    )
    # apply.ini AFTER add_argument: add_argument defaults overwrite set_defaults if called first
    if ini_defaults:
        parser.set_defaults(**ini_defaults)
    args = parser.parse_args()

    if not os.path.isdir(args.clients):
        print(f"ERROR: Client folder not found: {args.clients}", file=sys.stderr)
        sys.exit(1)

    if getattr(args, "bootstrap_client_risk", False):
        boot_out = args.bootstrap_output or args.client_ages
        if not boot_out:
            boot_out = os.path.join("data", "client_risk_pref.xlsx")
        boot_out = os.path.abspath(boot_out)
        print("\n=== Bootstrap client risk workbook ===")
        print(f"  Clients folder: {os.path.abspath(args.clients)}")
        print(f"  Output:         {boot_out}")
        try:
            n = build_client_risk_pref_workbook(
                args.clients, boot_out, merge=True
            )
        except ValueError as e:
            print(f"ERROR: {e}", file=sys.stderr)
            sys.exit(1)
        print(f"\n  Wrote {n} row(s). Open sheet 'ClientAges' and fill:")
        print("    • age — number")
        print("    • risk_preference — exactly one of: Averse, Moderate, Aggressive")
        print("\n  Next: set mfd_pack.ini (clients_folder, master, age_based=true,")
        print(f"         client_risk_pref or client_ages = path to this file), then run")
        print("         the exe without --bootstrap-client-risk.\n")
        sys.exit(0)

    master_path = args.master or default_master()
    out_xlsx = os.path.abspath(args.output)
    manifest_path = (
        args.manifest
        if args.manifest
        else os.path.splitext(out_xlsx)[0] + "_manifest.json"
    )

    if not os.path.isfile(master_path):
        print(f"ERROR: Master file not found: {master_path}", file=sys.stderr)
        sys.exit(1)

    if args.defensive_buckets:
        app_config.DEFENSIVE_BUCKETS = args.defensive_buckets

    defensive_used = app_config.DEFENSIVE_BUCKETS
    generated_ts = datetime.datetime.now(datetime.timezone.utc).replace(microsecond=0)

    print("\n=== MFD pack ===")
    print(f"  Config:     {ini_resolved or '(none — CLI / built-in defaults)'}")
    print(f"  Master:     {master_path}")
    print(f"  Clients:    {args.clients}")
    print(f"  Output:     {out_xlsx}")
    print(f"  Manifest:   {manifest_path}")
    if getattr(args, "age_based", False):
        cap = args.client_ages or app_config.CLIENT_AGES_FILE
        print(f"  Age-based:  ON  |  client_ages: {cap}")
        print(
            f"  Fallback archetype (if client missing from sidecar): {args.archetype}"
        )
    else:
        print(f"  Archetype:  {args.archetype} (fixed for all clients)")
    print(f"  Defensive:  {defensive_used}")
    print(f"  New cash:   {args.new_cash:,.0f} INR per client")
    if getattr(args, "allow_new_funds", False):
        fc = args.fund_choices or app_config.FUND_CHOICES_FILE
        print(f"  New funds:  ON  |  fund_choices: {fc}")
    else:
        print("  New funds:  OFF (existing holdings only for BUY pools)")

    print("\n  Loading master + running bulk_run (Format B)...")
    master_df = load_master_for_portfolio(master_path)

    fd, tmp_csv = tempfile.mkstemp(suffix=".csv", prefix="mfd_bulk_")
    os.close(fd)
    try:
        txn_df = bulk_run(
            format="B",
            client_dir=args.clients,
            master_df=master_df,
            archetype=args.archetype,
            new_cash=args.new_cash,
            output_csv=tmp_csv,
            age_based=getattr(args, "age_based", False),
            client_ages_path=getattr(args, "client_ages", None),
            allow_new_funds=getattr(args, "allow_new_funds", False),
            fund_choices_path=getattr(args, "fund_choices", None),
        )
    finally:
        try:
            os.unlink(tmp_csv)
        except OSError:
            pass

    n_txn = len(txn_df) if txn_df is not None and not txn_df.empty else 0

    print("\n  Loading holdings for Check Tab...")
    holdings_rows = load_all_clients(args.clients)
    if not holdings_rows:
        print("ERROR: No client holdings loaded from folder", file=sys.stderr)
        sys.exit(1)
    n_clients = len({r[0] for r in holdings_rows})

    wb = Workbook()
    wb.remove(wb.active)
    wb.create_sheet(INSTRUCTIONS_SHEET, 0)
    wb.create_sheet(PARAMETERS_SHEET, 1)
    wb.create_sheet(TRANSACTIONS_SHEET, 2)
    _write_instructions(wb[INSTRUCTIONS_SHEET])

    param_rows = [
        ("generated_at_utc", generated_ts.isoformat()),
        ("master_path", os.path.abspath(master_path)),
        ("clients_folder", os.path.abspath(args.clients)),
        ("input_format", "B (one file per client)"),
        (
            "age_based_glide_path",
            "yes" if getattr(args, "age_based", False) else "no",
        ),
        (
            "client_ages_workbook",
            (
                os.path.abspath(args.client_ages)
                if getattr(args, "client_ages", None)
                else (
                    os.path.abspath(app_config.CLIENT_AGES_FILE)
                    if getattr(args, "age_based", False)
                    else ""
                )
            ),
        ),
        ("archetype_parameter", args.archetype),
        (
            "target_policy_note",
            (
                "per-client glide (100 - k x age) from client_ages; missing clients use archetype_parameter"
                if getattr(args, "age_based", False)
                else "single ARCHETYPES target from archetype_parameter for all clients"
            ),
        ),
        ("new_cash_inr_per_client", str(args.new_cash)),
        ("defensive_buckets", defensive_used),
        (
            "allow_new_funds",
            "yes" if getattr(args, "allow_new_funds", False) else "no",
        ),
        (
            "fund_choices_workbook",
            (
                os.path.abspath(args.fund_choices)
                if getattr(args, "fund_choices", None)
                else (
                    os.path.abspath(app_config.FUND_CHOICES_FILE)
                    if getattr(args, "allow_new_funds", False)
                    else ""
                )
            ),
        ),
        (
            "rebalance_mode",
            (
                "existing_plus_fund_choices"
                if getattr(args, "allow_new_funds", False)
                else "existing_holdings_only"
            ),
        ),
        ("min_transaction_inr", str(app_config.MIN_TRANSACTION_AMT)),
        ("top_n_funds_per_direction", str(app_config.TOP_N_FUNDS)),
        ("output_xlsx", out_xlsx),
        ("clients_loaded", str(n_clients)),
        ("holding_lines", str(len(holdings_rows))),
        ("transaction_rows", str(n_txn)),
    ]
    _write_parameters(wb[PARAMETERS_SHEET], param_rows)
    _write_transactions(wb, txn_df if txn_df is not None else pd.DataFrame())

    master_read = master_path
    temp_master = None
    if _same_nav_path(master_path, LIVE_NAV_REPORTS):
        tfd, temp_master = tempfile.mkstemp(suffix=".xlsx")
        os.close(tfd)
        shutil.copyfile(master_path, temp_master)
        master_read = temp_master
        print("  NOTE: Master is live file - copied for read.")

    risk_lookup: dict[str, str] | None = None
    risk_formula: str | None = None
    try:
        print("  Copying Scheme types + latestNAV_Reports...")
        copy_nav_sheet(wb, master_read)
        print(f"  Writing {HOLDINGS_SHEET}...")
        client_ids = write_all_client_holdings(wb, holdings_rows)
        capath = getattr(args, "client_ages", None)
        if getattr(args, "age_based", False) and not capath:
            capath = app_config.CLIENT_AGES_FILE
        risk_lookup = _client_risk_for_check_tab(
            client_ids,
            capath,
            args.archetype,
        )
        tx_ws = wb[TRANSACTIONS_SHEET]
        risk_formula = risk_type_formula_from_transactions(
            tx_ws,
            archetype_fallback=args.archetype,
        )
        if risk_formula is None and risk_lookup:
            write_client_risk_map_sheet(wb, risk_lookup)
            lr = 1 + len(risk_lookup)
            risk_formula = risk_type_formula_from_client_risk_map(
                lr, archetype_fallback=args.archetype
            )
        print(f"  Writing {CHECK_SHEET}...")
        write_check_tab(
            wb,
            client_ids,
            nav_sheet_name=NAV_SHEET,
            risk_type_cell_formula=risk_formula,
            archetype_fallback=args.archetype,
        )
    finally:
        if temp_master and os.path.exists(temp_master):
            try:
                os.unlink(temp_master)
            except OSError:
                pass

    chk = wb[CHECK_SHEET]
    if risk_formula is None:
        chk["C5"].value = args.archetype
    chk["C6"].value = float(args.new_cash)

    sheet_order = [
        INSTRUCTIONS_SHEET,
        PARAMETERS_SHEET,
        TRANSACTIONS_SHEET,
        "Scheme types",
        NAV_SHEET,
        HOLDINGS_SHEET,
    ]
    if CLIENT_RISK_MAP_SHEET in wb.sheetnames:
        sheet_order.append(CLIENT_RISK_MAP_SHEET)
    sheet_order.append(CHECK_SHEET)
    _reorder_sheets(wb, sheet_order)
    wb.active = wb[INSTRUCTIONS_SHEET]

    os.makedirs(os.path.dirname(out_xlsx) or ".", exist_ok=True)
    wb.save(out_xlsx)
    print(f"\n  Saved workbook: {out_xlsx}")

    manifest = {
        "generated_at": generated_ts.isoformat(),
        "tool": "build_mfd_pack.py",
        "mfd_pack_ini": ini_resolved,
        "master_path": os.path.abspath(master_path),
        "clients_folder": os.path.abspath(args.clients),
        "input_format": "B",
        "archetype": args.archetype,
        "age_based": bool(getattr(args, "age_based", False)),
        "client_ages": (
            os.path.abspath(args.client_ages)
            if getattr(args, "client_ages", None)
            else (
                os.path.abspath(app_config.CLIENT_AGES_FILE)
                if getattr(args, "age_based", False)
                else None
            )
        ),
        "new_cash_inr": args.new_cash,
        "defensive_buckets": defensive_used,
        "allow_new_funds": bool(getattr(args, "allow_new_funds", False)),
        "fund_choices": (
            os.path.abspath(args.fund_choices)
            if getattr(args, "fund_choices", None)
            else (
                os.path.abspath(app_config.FUND_CHOICES_FILE)
                if getattr(args, "allow_new_funds", False)
                else None
            )
        ),
        "rebalance_mode": (
            "existing_plus_fund_choices"
            if getattr(args, "allow_new_funds", False)
            else "existing_holdings_only"
        ),
        "config_snapshot": {
            "MIN_TRANSACTION_AMT": app_config.MIN_TRANSACTION_AMT,
            "TOP_N_FUNDS": app_config.TOP_N_FUNDS,
            "OTHERS_MIN": app_config.OTHERS_MIN,
            "OTHERS_MAX_RATIO": app_config.OTHERS_MAX_RATIO,
            "ARCHETYPES": app_config.ARCHETYPES,
        },
        "output_xlsx": out_xlsx,
        "row_counts": {
            "clients": len(client_ids),
            "holding_lines": len(holdings_rows),
            "transactions": n_txn,
        },
    }
    with open(manifest_path, "w", encoding="utf-8") as f:
        json.dump(manifest, f, indent=2)

    print(f"  Manifest: {manifest_path}")
    print("\n=== Done ===\n")


if __name__ == "__main__":
    main()
